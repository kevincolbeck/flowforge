"""
FlowForge Test: Scrape website data into Excel
Scrapes top stories from Hacker News and saves to Excel
"""
import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def scrape_hacker_news():
    """Scrape top stories from Hacker News API"""
    print("Fetching top stories from Hacker News...")

    # Get top story IDs
    response = httpx.get("https://hacker-news.firebaseio.com/v0/topstories.json")
    story_ids = response.json()[:20]  # Get top 20

    stories = []
    for i, story_id in enumerate(story_ids):
        print(f"  Fetching story {i+1}/20...", end="\r")
        story_response = httpx.get(f"https://hacker-news.firebaseio.com/v0/item/{story_id}.json")
        story = story_response.json()
        stories.append({
            "rank": i + 1,
            "title": story.get("title", "N/A"),
            "url": story.get("url", f"https://news.ycombinator.com/item?id={story_id}"),
            "score": story.get("score", 0),
            "author": story.get("by", "unknown"),
            "comments": story.get("descendants", 0),
            "time": datetime.fromtimestamp(story.get("time", 0)).strftime("%Y-%m-%d %H:%M")
        })

    print("\n  Done fetching stories!")
    return stories

def save_to_excel(data, filename):
    """Save scraped data to Excel with formatting"""
    print(f"Creating Excel file: {filename}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Hacker News Top Stories"

    # Headers
    headers = ["Rank", "Title", "URL", "Score", "Author", "Comments", "Posted"]
    header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row, story in enumerate(data, 2):
        ws.cell(row=row, column=1, value=story["rank"])
        ws.cell(row=row, column=2, value=story["title"])
        ws.cell(row=row, column=3, value=story["url"])
        ws.cell(row=row, column=4, value=story["score"])
        ws.cell(row=row, column=5, value=story["author"])
        ws.cell(row=row, column=6, value=story["comments"])
        ws.cell(row=row, column=7, value=story["time"])

    # Adjust column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 18

    wb.save(filename)
    print(f"Saved {len(data)} stories to {filename}")

if __name__ == "__main__":
    print("=" * 50)
    print("FlowForge Test: Web Scraping to Excel")
    print("=" * 50)

    # Scrape data
    stories = scrape_hacker_news()

    # Save to Excel
    output_file = "C:/Users/kevin/hacker_news_top_stories.xlsx"
    save_to_excel(stories, output_file)

    print("\nDone! Open the Excel file to see results.")
    print(f"File: {output_file}")
